# Importando bibliotecas
import sys
import os
import pandas as pd
from openpyxl import load_workbook, Workbook


# Alterando o diretório para './input' e salvando os arquivos em lista
def open_repository(path):
    os.chdir(path)
    return os.listdir()


# Verificando se os arquivos são de extensão suportada
def is_excel_file(files):
    supported_files = ['.xls', '.xlsx', '.xlsb']

    for file in files:
        file_type = os.path.splitext(file)[1]
        if file_type not in supported_files:
            return False
    
    return True


# Verificando se os arquivos são de extensão xlsb
def is_xlsb(files):
    xlsb_files = []

    for file in files:
        file_type = os.path.splitext(file)[1]
        if file_type == '.xlsb':
            xlsb_files.append(file)
    
    return xlsb_files if xlsb_files else False


# Convertendo os arquivos em xlsb para xlsx e removendo-os do diretório
def xlsb_convertion(xlsb_files):
    df = pd.read_excel(xlsb_files[0], sheet_name=None, engine="pyxlsb")
    sheets = list(df.keys())
    print("")
    
    # Retorna as abas da planilha para o usuário selecionar qual será copiada
    print("Sheets:")
    for i, sheet in enumerate(sheets):
        print(str(i + 1) + '.' + sheet)
    print("")

    # Solicita que usuário selecione a aba que será copiada e retorna erro caso não seja digitado um número
    try: 
        sheet = int(input("Choose the sheet to copy (number): "))
    except ValueError: 
        sys.exit("Only numbers accepted")
    print("")
    
    # Retorna erro caso o usuário selecione uma aba que não está na lista
    if sheet > len(sheets) or sheet < 1:
        sys.exit("Error: Sheet out of range")
    
    sheet_title = sheets[sheet-1]
    
    # Converte os arquivos para xlsx
    for file in xlsb_files:
        df = pd.read_excel(file, sheet_name=sheets[sheet-1], engine="pyxlsb")
        df.to_excel(os.path.splitext(file)[0] + ".xlsx", sheet_name=sheet_title, index=False)
    
    # Remove os arquivos xlsb 
    for file in xlsb_files:
        os.remove(file)
    
    # Instancia o primeiro arquivo da lista como Workbook de Openpyxl
    files = os.listdir()
    wb_in = load_workbook(filename = files[0])
    ws_in = wb_in.active
    
    return ws_in, sheet_title


# Carregando o primeiro arquivo da pasta ./input e imprimindo as abas para definição do usuário
def define_sheet(files):
    # Instancia o primeiro arquivo da lista como Workbook de Openpyxl
    wb_in = load_workbook(filename = files[0])
    print("")
    
    # Retorna as abas da planilha para o usuário selecionar qual será copiada
    print("Sheets:")
    for i, sheet in enumerate(wb_in):
        print(str(i + 1) + '.' + sheet.title)
    print("")
    
    # Solicita que usuário selecione a aba que será copiada e retorna erro caso não seja digitado um número
    try: 
        sheet = int(input("Choose the sheet to copy (number): "))
    except ValueError:
        sys.exit("Only numbers accepted")
    print("")

    # Retorna erro caso o usuário selecione uma aba que não está na lista
    if sheet > len(wb_in.sheetnames) or sheet < 1:
        sys.exit("Error: Sheet out of range")
        
    sheet_title = wb_in.sheetnames[sheet-1]
    ws_in = wb_in[sheet_title]
    
    return ws_in, sheet_title


def define_header(worksheet):
    # Solicita que usuário informe a linha que se encontra o cabeçalho e retorna erro caso não seja digitado um número
    try:
        header_row = int(input("Header row (number): "))
    except ValueError:
        sys.exit("Only numbers accepted")
        
    header = []

    for cell in worksheet[header_row]:
        header.append(cell.value)

    # Retornando erro se a linha informada estiver vazia
    if header == [None]:
        sys.exit("Header is empty")
    
    return header, header_row


def define_columns(ws_in, header):
    print("")
    # Retorna as colunas para o usuário selecionar quais serão copiada
    print("Columns:")
    for i, title in enumerate(header):
        print(str(i + 1) + '.' + str(title))

    # Adiciona ao final uma coluna 'file' para trazer a informação do arquivo
    header.append('File')

    # Solicita que o usuário informe as colunas que devem ser copiadas e retorna erro caso não seja digitado um número
    print("")
    target_cols = input("Choose the columns to copy (numbers separated by comma): ")
    columns = target_cols.split(',')
    columns.append(header.index('File')+1)

    try:
        columns = list(map(lambda x: int(x), columns))
    except ValueError:
        sys.exit("Only number accepted")

    # Retorna erro se a coluna não existir
    for column in columns:
        if column > (ws_in.max_column + 1): # +1 pois foi adicionado no header a coluna 'File'
            sys.exit("Error: Column out of range")

    return columns


def create_workbook(header, columns):
    #Criando novo Workbook para colar dados copiados
    wb_out = Workbook()
    ws_out = wb_out.active

    # Adicionando o cabeçalho no arquivo criado
    for i, title in enumerate(columns):
        ws_out.cell(row=1, column=i+1).value = header[title - 1]

    return wb_out, ws_out


def consolidate_data(sheet_title, columns, header_row, ws_out):
    files = os.listdir()
    
    # Loop para cada arquivo
    for file in files:
        wb_in = load_workbook(filename = file)
        
        # Verifica se a aba procurada existe no arquivo, se não retorna erro
        if not sheet_title in wb_in.sheetnames:
            sys.exit(f"Sheet '{sheet_title}' does not exist in file '{file}'")
        
        ws_in = wb_in[sheet_title]
        
        # Iteração do input dos dados e inclusão em lista
        data = []
        for i in range(1, len(columns)):
            col_data = []
            for j in range(header_row+1, ws_in.max_row+1):
                value = ws_in.cell(row=j, column=int(columns[i-1])).value
                col_data.append(value)
                    
            data.append(col_data)
        
        # Adicionando coluna com dados do arquivo
        file_column = []
        
        for i in range(1, len(data[0]) + 1):
            file_column.append(file)    
        
        data.append(file_column)
        
        # Iteração do output dos dados em novo arquivo
        last_row_out = ws_out.max_row
        
        for i, column in enumerate(data):
            for j, cell in enumerate(column):
                ws_out.cell(row=(j+last_row_out+1), column=(i+1)).value = cell


def save_workbook(out_workbook):
    os.chdir('../output')
    out_workbook.save('output.xlsx')
    print('Successfully completed')


def main():
    files = open_repository('./input')
    
    is_excel = is_excel_file(files)
    
    if not is_excel:
        sys.exit("Error: Only Excel file supported")
        
    xlsb_files = is_xlsb(files)
    
    in_worksheet, sheet_title = xlsb_convertion(xlsb_files) if xlsb_files else define_sheet(files)
    
    header, header_row = define_header(in_worksheet)
    columns = define_columns(in_worksheet, header)
    out_workbook, out_worksheet = create_workbook(header, columns)
    consolidate_data(sheet_title, columns, header_row, out_worksheet)
    save_workbook(out_workbook)


if __name__ == "__main__":
    main()